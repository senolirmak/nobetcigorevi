#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun May 11 11:08:09 2025

@author: senolirmak
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from PyQt5.QtCore import QDateTime
from utils.database_util import TeacherManager

import os
import subprocess
from pathlib import Path

def create_excel_report(sonuc=None):
    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Nöbet Dağıtım Raporu"
    data = TeacherManager()
    
    # Get current date in Turkish format
    tarih = QDateTime.currentDateTime().toString("dddd - dd.MM.yyyy")
    file_tarih_saat = QDateTime.currentDateTime().toString("ddMMyyyyHHmm")
    file_path = "raporlar/"
    output_filename = f"{file_path}Rapor_{file_tarih_saat}.xlsx"

    sonuc['assignments'] = sorted(sonuc['assignments'], key=lambda x: (x['teacher_id'], x['hour']))
    # Main title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = f"{tarih} DERS DOLDURMA GÖREVLERİ RAPORU"
    title_cell.font = Font(size=12, bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add borders
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
            
    # Column headers
    headers = ["Devamsız Öğretmen", "Saat", "Sınıf", "Nöbetçi Öğretmen"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Add assignment data
    row_num = 3
    for atama in sonuc['assignments']:
        ws.cell(row=row_num, column=2, value=f"{atama['hour']}. Ders").alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=3, value=atama['class']).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=1, value=data.get_ogretmen_adi(atama['absent_teacher_id']))
        ws.cell(row=row_num, column=4, value=data.get_ogretmen_adi(atama['teacher_id']))
        row_num += 1
    
    for row in ws.iter_rows(min_row=2, max_row=row_num-1, max_col=4):
        for cell in row:
            cell.border = thin_border
          
    # Unassigned lessons section
    row_num += 2  # Add extra space before Unassigned lessons
    
    if sonuc['unassigned']:
        ws.cell(row=row_num, column=1, value="ATANAMAYAN DERSLER").font = Font(bold=True)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        row_num += 1
        
        headers = ["Devamsız Öğretmen", "Saat", "Sınıf", "Nöbetçi Öğretmen"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        row_num += 1
        for ders in sonuc['unassigned']:
            ws.cell(row=row_num, column=2, value=f"{ders['hour']}. Ders").alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=3, value=ders['class']).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=1, value=data.get_ogretmen_adi(ders['absent_teacher_id']))
            ws.cell(row=row_num, column=4, value="-----")
            row_num += 1
    
    # Teacher assignment statistics
    row_num += 2  # Add extra space before statistics

    # Statistics title
    ws.cell(row=row_num, column=1, value="NÖBETÇİ DAĞILIM İSTATİSTİKLERİ").font = Font(bold=True)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    row_num += 1
    
    # Statistics headers
    stats_headers = ["Öğretmen Adı", "Toplam Ders Doldurma Sayısı"]
    for col_num, header in enumerate(stats_headers, 1):  # Start from column 3
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Statistics data
    row_num += 1
    for teacher_id, count in sonuc['teacher_counts'].items():
        teacher_name = data.get_ogretmen_adi(teacher_id)
        ws.cell(row=row_num, column=1, value=teacher_name)
        ws.cell(row=row_num, column=2, value=count).alignment = Alignment(horizontal='center')
        row_num += 1

    
    # 1. Devamsız öğretmenlerin toplam ders saatlerini hesapla
        # Sonuçları saklamak için sözlükler
    atanan_saatler = defaultdict(int)
    atanamayan_saatler = defaultdict(int)
    
    # Atanan ders saatlerini hesapla
    for atama in sonuc['assignments']:
        devamsiz_id = atama['absent_teacher_id']
        atanan_saatler[devamsiz_id] += 1
    
    # Atanamayan ders saatlerini hesapla
    for atama in sonuc['unassigned']:
        devamsiz_id = atama['absent_teacher_id']
        atanamayan_saatler[devamsiz_id] += 1
    
    # Tüm devamsız öğretmenleri birleştir
    tum_devamsizlar = set(atanan_saatler.keys()) | set(atanamayan_saatler.keys())
    temp_dict = {}
    temp_list =[]
    for devamsiz_id in sorted(tum_devamsizlar):
        temp_dict['devamsiz_adi'] = data.get_ogretmen_adi(devamsiz_id)
        temp_dict['atanan'] = atanan_saatler.get(devamsiz_id, 0)
        temp_dict['atanamayan'] = atanamayan_saatler.get(devamsiz_id, 0)
        temp_dict['toplam'] = atanan_saatler.get(devamsiz_id, 0) + atanamayan_saatler.get(devamsiz_id, 0)
        ds = temp_dict.copy()
        temp_list.append(ds)
        temp_dict.clear()
        
    row_num += 2  # Add extra space before statistics

    # Statistics title
    ws.cell(row=row_num, column=1, value="DEVAMSIZ ÖĞRETMENLERİN TOPLAM DERS SAATİ").font = Font(bold=True)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    row_num += 1
    
    # Statistics Devamsız Öğretmen Ders Saati

    stats_headers = ["Devamsız Öğretmen", "Atanan","Atanamayan", "Toplam" ]
    for col_num, header in enumerate(stats_headers, 1):  # Start from column 3
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    row_num += 1    
    for devamsiz in temp_list:
        ws.cell(row=row_num, column=1, value=devamsiz['devamsiz_adi'])
        ws.cell(row=row_num, column=2, value=devamsiz['atanan']).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=3, value=devamsiz['atanamayan']).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=4, value=devamsiz['toplam']).alignment = Alignment(horizontal='center')
        row_num += 1
    
    # Adjust column widths
    column_widths = [25, 12, 15, 25]  # Adjusted widths for better display
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    
    
    # Save file
    wb.save(output_filename)
    print(f"Rapor başarıyla oluşturuldu: {output_filename}")
    # Kullanım
    open_excel_file(output_filename)
    

def find_teacher_name(teacher_id, atamalar):
    for atama in atamalar:
        if atama['nobetci_id'] == teacher_id:
            return atama['nobetci_adi']
    return "Bilinmiyor"

def open_excel_file(file_path):
    """Belirtilen Excel dosyasını sistemdeki varsayılan uygulamayla açar."""
    path = Path(file_path).resolve()
    if not path.exists():
        print(f"❌ Dosya bulunamadı: {path}")
        return

    try:
        if os.name == "posix":  # Linux / macOS
            subprocess.Popen(["xdg-open", str(path)])
        elif os.name == "nt":  # Windows
            os.startfile(str(path))
        else:
            subprocess.Popen(["open", str(path)])  # macOS alternatif
        print(f"📂 '{path.name}' dosyası açılıyor...")
    except Exception as e:
        print(f"⚠️ Dosya açılamadı: {e}")

