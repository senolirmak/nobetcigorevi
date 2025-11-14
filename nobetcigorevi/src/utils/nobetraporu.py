#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel tabanlı Nöbetçi Öğretmen Ders Doldurma Raporu oluşturucu.
@author: Şenol Irmak
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import QDateTime
from collections import defaultdict
from utils.database_util import TeacherManager
from pathlib import Path
import os
import subprocess

from openpyxl.worksheet.page import PageMargins
from datetime import datetime, timedelta
from db.database import SessionLocal
from db.models import NobetGorevi, NobetOgretmen
from sqlalchemy import func



class ExcelRaporOlusturucu:
    def __init__(self, hedef_klasor="raporlar"):
        """Rapor oluşturucu başlatılır."""
        self.data = TeacherManager()
        self.hedef_klasor = hedef_klasor
        Path(hedef_klasor).mkdir(parents=True, exist_ok=True)

    # ---------------------------------------------------------
    # 🔹 Ana Rapor Fonksiyonu
    # ---------------------------------------------------------
    def create_excel_report(self, sonuc):
        """Ders doldurma raporunu oluşturur ve açar."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Nöbet Dağıtım Raporu"

        tarih = QDateTime.currentDateTime().toString("dddd - dd.MM.yyyy")
        file_tarih_saat = QDateTime.currentDateTime().toString("ddMMyyyyHHmm")
        output_filename = f"{self.hedef_klasor}/Rapor_{file_tarih_saat}.xlsx"

        sonuc['assignments'] = sorted(sonuc['assignments'], key=lambda x: (x['teacher_id'], x['hour']))

        # 🔸 Başlık
        ws.merge_cells('A1:D1')
        ws['A1'].value = f"{tarih} DERS DOLDURMA GÖREVLERİ RAPORU"
        ws['A1'].font = Font(size=12, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # 🔸 Sütun başlıkları
        headers = ["Devamsız Öğretmen", "Saat", "Sınıf", "Nöbetçi Öğretmen"]
        self._add_headers(ws, 2, headers)

        # 🔸 Dağılım verisi
        row_num = 3
        for atama in sonuc['assignments']:
            ws.cell(row=row_num, column=1, value=self.data.get_ogretmen_adi(atama['absent_teacher_id']))
            ws.cell(row=row_num, column=2, value=f"{atama['hour']}. Ders").alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=3, value=atama['class']).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=4, value=self.data.get_ogretmen_adi(atama['teacher_id']))
            row_num += 1

        # Kenarlık uygula
        self._apply_borders(ws, 2, row_num - 1, 4)

        # 🔸 Atanamayan dersler
        if sonuc['unassigned']:
            row_num += 2
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=1, value="ATANAMAYAN DERSLER").font = Font(bold=True)
            row_num += 1
            self._add_headers(ws, row_num, headers)
            row_num += 1

            for ders in sonuc['unassigned']:
                ws.cell(row=row_num, column=1, value=self.data.get_ogretmen_adi(ders['absent_teacher_id']))
                ws.cell(row=row_num, column=2, value=f"{ders['hour']}. Ders").alignment = Alignment(horizontal='center')
                ws.cell(row=row_num, column=3, value=ders['class']).alignment = Alignment(horizontal='center')
                ws.cell(row=row_num, column=4, value="-----").alignment = Alignment(horizontal='center')
                row_num += 1

        # 🔸 Öğretmen bazlı istatistikler
        row_num = self._add_teacher_stats(ws, sonuc, row_num + 2)

        # 🔸 Devamsız öğretmen saatleri
        self._add_absent_teacher_stats(ws, sonuc, row_num + 2)

        # 🔸 Sütun genişlikleri
        self._set_column_widths(ws, [25, 12, 15, 25])

        # 🔸 Kaydet ve aç
        wb.save(output_filename)
        print(f"✅ Rapor başarıyla oluşturuldu: {output_filename}")
        self.open_excel_file(output_filename)
    
    def raporla_nobet_gorevi_excel(self, uygulama_tarihi_str=None, hedef_klasor="raporlar"):
        """
        NobetGorevi tablosundaki en güncel (veya verilen) uygulama_tarihi'ne göre
        Türkçe gün adlarıyla, Pazartesi→Cuma sıralı ve 'Nöbet Günü' ilk sütunda olacak şekilde
        Excel raporu oluşturur.
        """
        session = SessionLocal()
        try:
            # 🔹 1. Tarih belirle (label'dan veya DB'den)
            if not uygulama_tarihi_str or uygulama_tarihi_str.strip() == "":
                latest_date = session.query(func.max(NobetGorevi.uygulama_tarihi)).scalar()
                if not latest_date:
                    raise ValueError("Veritabanında nöbet kaydı bulunamadı.")
                uygulama_tarihi = latest_date.date()
            else:
                try:
                    uygulama_tarihi = datetime.strptime(uygulama_tarihi_str, "%d.%m.%Y").date()
                except ValueError:
                    uygulama_tarihi = datetime.strptime(uygulama_tarihi_str, "%Y-%m-%d").date()
            
            simdi = datetime.now().date()
            hafta_baslangici = simdi - timedelta(days=simdi.weekday())  # Pazartesi
            hafta_no = hafta_baslangici.isocalendar()[1]
            simdi_ay = simdi.month
            
            # Dönem sözlüğü
            donem_sayi = {
                1: [9, 10, 11, 12, 1],  # 1. dönem ayları
                2: [2, 3, 4, 5, 6]      # 2. dönem ayları
            }
            
            # Dönem numarasını belirle
            donem_numarasi = next((k for k, v in donem_sayi.items() if simdi_ay in v), None)
            donem = f"{donem_numarasi}. Dönem" if donem_numarasi else "Yaz Dönemi"

            # 🔹 2. Gün çevirim haritası (TR ↔ EN)
            gun_map = {
                "Monday": "Pazartesi",
                "Tuesday": "Salı",
                "Wednesday": "Çarşamba",
                "Thursday": "Perşembe",
                "Friday": "Cuma",
                "Saturday": "Cumartesi",
                "Sunday": "Pazar"
            }

            # 🔹 3. Gün sırası (Pzt→Cuma)
            gun_sirasi = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

            # 🔹 4. Veritabanından nöbet kayıtlarını çek
            kayitlar = (
                session.query(NobetGorevi)
                .join(NobetOgretmen, NobetGorevi.ogretmen_id == NobetOgretmen.id)
                .filter(func.date(NobetGorevi.uygulama_tarihi) == uygulama_tarihi)
                .all()
            )

            if not kayitlar:
                print(f"❗ {uygulama_tarihi} tarihli nöbet kaydı bulunamadı.")
                return None

            # 🔹 5. Günlere göre grupla
            grouped = {g: [] for g in gun_sirasi}  # sadece hafta içi günler
            for g in kayitlar:
                gun_tr = gun_map.get(g.nobet_gun, g.nobet_gun)
                if gun_tr in grouped:
                    grouped[gun_tr].append(g)

            # 🔹 6. Excel oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Nöbetçi Öğretmenler"
            
            # Sayfa ayarları (A4, yatay, tek sayfa)
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # Dikey
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.page_margins = PageMargins(
                left=1, right=0.3, top=0.5, bottom=0.5, header=0.2, footer=0.2
            )

            # Başlıklar
            ws.merge_cells("A1:C1")
            ws["A1"] = "Abdurrahim Karakoç Anadolu Lisesi"
            ws["A1"].font = Font(size=14, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:C2")
            ws["A2"] = f"Nöbetçi Öğretmen Listesi – Uygulama Tarihi: {uygulama_tarihi.strftime('%d.%m.%Y')}"
            ws["A2"].font = Font(size=12, italic=True)
            ws["A2"].alignment = Alignment(horizontal="center")
            
            ws.merge_cells("A3:C3")
            ws["A3"] = f"{donem} – {hafta_no}. Hafta"
            ws["A3"].font = Font(size=12,italic=True ,bold=False)
            ws["A3"].alignment = Alignment(horizontal="center")

            ws.append(["", "", "", ""])  # Boş satır
            current_row = 5

            # Biçimlendirme stilleri
            header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            thin = Side(border_style="thin", color="000000")

            # 🔹 7. Günlere göre sıralı tablo üretimi
            for gun in gun_sirasi:
                ogretmenler = grouped.get(gun, [])
                if not ogretmenler:
                    continue

                # Gün başlığı
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                ws.cell(row=current_row, column=1).value = f"{gun} Günü Nöbetçileri"
                ws.cell(row=current_row, column=1).font = Font(size=12, bold=True, color="1F497D")
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left")
                current_row += 1

                # Sütun başlıkları
                headers = ["Öğretmen Adı", "Branş", "Nöbet Yeri"]
                ws.append(headers)
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                current_row += 1

                # Kayıtlar
                for g in ogretmenler:
                    ws.append([
                        g.ogretmen.adi_soyadi,
                        g.ogretmen.brans,
                        g.nobet_yeri
                    ])
                    for col in range(1, 4):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        cell.alignment = Alignment(horizontal="left")
                    current_row += 1

                # Gün arası boşluk
                ws.append(["", "", "", ""])
                current_row += 1

            # 🔹 8. Sütun genişlikleri
            #ws.column_dimensions["A"].width = 16  # Nöbet günü
            ws.column_dimensions["A"].width = 28  # Öğretmen
            ws.column_dimensions["B"].width = 28  # Branş
            ws.column_dimensions["C"].width = 28  # Nöbet yeri

            # 🔹 9. Dosyayı kaydet
            if not os.path.exists(hedef_klasor):
                os.makedirs(hedef_klasor)

            rapor_adi = f"Rapor_Nobet_{uygulama_tarihi.strftime('%Y%m%d')}.xlsx"
            rapor_yolu = os.path.join(hedef_klasor, rapor_adi)
            wb.save(rapor_yolu)

            print(f"✅ Nöbet raporu oluşturuldu: {rapor_yolu}")
            self.open_excel_file(rapor_yolu)
            
        except Exception as e:
            print(f"❌ Rapor oluşturulamadı: {e}")
            return None
        finally:
            session.close()

    # ---------------------------------------------------------
    # 🔹 Yardımcı Fonksiyonlar
    # ---------------------------------------------------------
    def _add_headers(self, ws, row_num, headers):
        """Tablo başlıklarını ekler."""
        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = fill

    def _apply_borders(self, ws, min_row, max_row, max_col):
        """Hücre kenarlıkları ekler."""
        thin = Side(style='thin')
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def _set_column_widths(self, ws, widths):
        """Sütun genişliklerini ayarla."""
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _add_teacher_stats(self, ws, sonuc, row_num):
        """Nöbetçi öğretmen istatistiklerini tabloya ekler."""
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        ws.cell(row=row_num, column=1, value="NÖBETÇİ DAĞILIM İSTATİSTİKLERİ").font = Font(bold=True)
        row_num += 1

        headers = ["Öğretmen Adı", "Toplam Ders Doldurma Sayısı"]
        self._add_headers(ws, row_num, headers)
        row_num += 1

        for teacher_id, count in sonuc['teacher_counts'].items():
            ws.cell(row=row_num, column=1, value=self.data.get_ogretmen_adi(teacher_id))
            ws.cell(row=row_num, column=2, value=count).alignment = Alignment(horizontal='center')
            row_num += 1
        return row_num

    def _add_absent_teacher_stats(self, ws, sonuc, row_num):
        """Devamsız öğretmenlerin toplam ders saati istatistiği ekler."""
        atanan_saatler = defaultdict(int)
        atanamayan_saatler = defaultdict(int)

        for a in sonuc['assignments']:
            atanan_saatler[a['absent_teacher_id']] += 1
        for a in sonuc['unassigned']:
            atanamayan_saatler[a['absent_teacher_id']] += 1

        tum_devamsizlar = set(atanan_saatler.keys()) | set(atanamayan_saatler.keys())
        tablo = []
        for devamsiz_id in sorted(tum_devamsizlar):
            tablo.append({
                "devamsiz_adi": self.data.get_ogretmen_adi(devamsiz_id),
                "atanan": atanan_saatler.get(devamsiz_id, 0),
                "atanamayan": atanamayan_saatler.get(devamsiz_id, 0),
                "toplam": atanan_saatler.get(devamsiz_id, 0) + atanamayan_saatler.get(devamsiz_id, 0)
            })

        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        ws.cell(row=row_num, column=1, value="DEVAMSIZ ÖĞRETMENLERİN TOPLAM DERS SAATİ").font = Font(bold=True)
        row_num += 1

        headers = ["Devamsız Öğretmen", "Atanan", "Atanamayan", "Toplam"]
        self._add_headers(ws, row_num, headers)
        row_num += 1

        for d in tablo:
            ws.cell(row=row_num, column=1, value=d["devamsiz_adi"])
            ws.cell(row=row_num, column=2, value=d["atanan"]).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=3, value=d["atanamayan"]).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=4, value=d["toplam"]).alignment = Alignment(horizontal='center')
            row_num += 1

    # ---------------------------------------------------------
    # 🔹 Excel Dosyası Açma
    # ---------------------------------------------------------
    def open_excel_file(self, file_path):
        """Excel dosyasını sistemde varsayılan uygulamayla açar."""
        path = Path(file_path).resolve()
        if not path.exists():
            print(f"❌ Dosya bulunamadı: {path}")
            return

        try:
            if os.name == "posix":  # Linux / macOS
                subprocess.Popen(["xdg-open", str(path)])
            elif os.name == "nt":  # Windows
                os.startfile(str(path))
            else:
                subprocess.Popen(["open", str(path)])
            print(f"📂 '{path.name}' dosyası açılıyor...")
        except Exception as e:
            print(f"⚠️ Dosya açılamadı: {e}")

